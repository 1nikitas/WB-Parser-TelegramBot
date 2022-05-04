
# -*- coding: cp1251 -*-
# import aiogram
import os
import openpyxl
from aiogram import Bot, Dispatcher, types, executor
from aiogram.dispatcher.filters import Command
from selenium import webdriver
from selenium.webdriver.common.by import By
from tqdm import tqdm

def parser(url):
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    browser = webdriver.Chrome('chromedriver.exe', options=options)
    browser.get(url)
    try:
        product = browser.find_element(By.TAG_NAME, 'h1')
    except Exception:
        product = ""

    try:
        price = browser.find_element(By.CLASS_NAME, 'new-price')
    except Exception:
        price = ""

    try:
        old_price = browser.find_element(By.CLASS_NAME, 'price-offline__old-price')
    except Exception:
        old_price = price

    product = product.text
    price1 = price.text.split(" ")[0]
    old_price1 = old_price.text.split(" ")[0]
    diff = int(old_price1) - int(price1)
    return product, price1, old_price1, diff


TOKEN = "1674260748:AAFfDDD6-_kZIuHEOosch3cx5Sa5tO9UTHY"
bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
admins_ids = [582897416, 421048811]
@dp.message_handler(content_types=['document'])
async def doc(message: types.Message):
    if message.from_user.id in admins_ids:
        await message.answer("Все ок!")
        doc_id = message.document.file_id
        file_info = await bot.get_file(doc_id)
        path = file_info.file_path
        await message.document.download(path)
    else:
        await message.answer("В доступе отказано!")

@dp.message_handler(Command('getFile'))
async def upload(message: types.Message):
    if message.from_user.id in admins_ids:
        msg = await message.answer("Началась обработка запроса!")
        path = r'C:\Users\nkise\PycharmProjects\Pars_Prac\documents'
        files = os.listdir(path)
        file = min(files)
        path = path + "/" + str(file)
        exel_file = openpyxl.load_workbook(path)
        sheet = exel_file.active

        for row in range(2, sheet.max_row+1):
            name, price, old_price, discount = list(parser(sheet.cell(row=row, column=1).value)) #product, price1, old_price1, diff

            sheet.cell(row=row, column=3).value = old_price
            sheet.cell(row=row, column=4).value = price
            sheet.cell(row=row, column=5).value = discount
            await msg.edit_text(f'Progress : {int((100*row/(sheet.max_row+1)))} %')
        await msg.edit_text("Progress : 100 %")
        file_name = 'ready.xlsx'
        exel_file.save(f'{file_name}')
        await message.reply_document(open(f'{file_name}', 'rb'))

        exel_file.close()
    else:
        await message.answer("В доступе отказано!")



if __name__ == "__main__":
    executor.start_polling(dp,skip_updates=True )